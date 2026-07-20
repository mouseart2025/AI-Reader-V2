"""Series Bible export_all 全量导出回归测试 (issue #39).

默认保持 Top-N 截断(向后兼容);export_all=True 时输出全部实体/关系/事件。
"""

from openpyxl import load_workbook

from src.services.series_bible_renderer import render_markdown
from src.services.series_bible_service import SeriesBibleData
from src.services.xlsx_renderer import render_xlsx


def _make_data(n_edges: int = 80, n_events: int = 1200) -> SeriesBibleData:
    return SeriesBibleData(
        novel_title="测试小说",
        novel_author=None,
        chapter_range=(1, 100),
        modules=["relations", "timeline"],
        relations={
            "edges": [
                {
                    "source": f"人物{i}",
                    "target": f"人物{i + 1}",
                    "relation_type": "朋友",
                    "weight": 100 - i,
                }
                for i in range(n_edges)
            ]
        },
        timeline=[
            {"chapter": 1, "summary": f"事件{i}", "type": "战斗", "importance": "mid"}
            for i in range(n_events)
        ],
    )


def _md_relation_rows(md: str) -> list[str]:
    # 排除表头 "| 人物A | 人物B |..."，只数数据行
    return [ln for ln in md.splitlines() if ln.startswith("| 人物") and not ln.startswith("| 人物A")]


def test_markdown_relations_capped_by_default():
    md = render_markdown(_make_data(), export_all=False)
    assert len(_md_relation_rows(md)) == 30  # 默认 top-30


def test_markdown_relations_export_all():
    md = render_markdown(_make_data(), export_all=True)
    assert len(_md_relation_rows(md)) == 80


def test_xlsx_relations_capped_by_default():
    buf = render_xlsx(_make_data(), export_all=False)
    ws = load_workbook(buf)["关系表"]
    assert ws.max_row == 1 + 50  # header + top-50


def test_xlsx_relations_export_all():
    buf = render_xlsx(_make_data(), export_all=True)
    ws = load_workbook(buf)["关系表"]
    assert ws.max_row == 1 + 80


def test_xlsx_timeline_capped_by_default():
    buf = render_xlsx(_make_data(), export_all=False)
    ws = load_workbook(buf)["时间线"]
    assert ws.max_row == 1 + 1000  # header + top-1000


def test_xlsx_timeline_export_all():
    buf = render_xlsx(_make_data(), export_all=True)
    ws = load_workbook(buf)["时间线"]
    assert ws.max_row == 1 + 1200
