"""Series Bible Excel (.xlsx) renderer — convert collected data to multi-sheet workbook.

Uses openpyxl to generate a styled .xlsx with one sheet per module:
Characters, Relations, Locations, Items, Orgs, Timeline.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.services.series_bible_service import SeriesBibleData

# ── Shared styles ──────────────────────────────────

_HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(vertical="top", wrap_text=True)


def render_xlsx(data: SeriesBibleData, template: str = "complete") -> io.BytesIO:
    """Render SeriesBibleData as an Excel workbook. Returns BytesIO buffer."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    modules = data.modules or [
        "characters", "relations", "locations", "items", "orgs", "timeline",
    ]

    if "characters" in modules and data.characters:
        _render_characters(wb, data.characters)

    if "relations" in modules and data.relations:
        _render_relations(wb, data.relations)

    if "locations" in modules and data.locations:
        _render_locations(wb, data.locations)

    if "items" in modules and data.items:
        _render_items(wb, data.items)

    if "orgs" in modules and data.orgs:
        _render_orgs(wb, data.orgs)

    if "timeline" in modules and data.timeline:
        _render_timeline(wb, data.timeline)

    # Ensure at least one sheet exists
    if len(wb.sheetnames) == 0:
        wb.create_sheet("空")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Sheet renderers ────────────────────────────────


def _render_characters(wb: Workbook, characters: list[dict]) -> None:
    ws = wb.create_sheet("角色表")
    headers = ["名称", "别称", "首次登场", "出场章数", "能力", "关系", "主要经历"]
    _write_header(ws, headers)

    for ch in characters:
        aliases = ch.get("aliases", [])
        alias_names = [a["name"] for a in aliases if a["name"] != ch["name"]]

        abilities = ch.get("abilities", [])
        ab_str = "; ".join(
            f"{a.get('dimension', '')}·{a.get('name', '')}" for a in abilities[:5]
        )

        relations = ch.get("relations", [])
        rel_parts = []
        for rel in relations[:5]:
            other = rel.get("other_person", "")
            stages = rel.get("stages", [])
            rel_type = stages[-1].get("relation_type", "") if stages else ""
            rel_parts.append(f"{other}({rel_type})" if rel_type else other)

        experiences = ch.get("experiences", [])
        exp_str = "; ".join(e.get("summary", "")[:40] for e in experiences[:3])

        stats = ch.get("stats", {})

        ws.append([
            ch["name"],
            ", ".join(alias_names),
            stats.get("first_chapter", ""),
            stats.get("chapter_count", 0),
            ab_str,
            ", ".join(rel_parts),
            exp_str,
        ])

    _auto_width(ws, headers)


def _render_relations(wb: Workbook, relations: dict) -> None:
    ws = wb.create_sheet("关系表")
    headers = ["人物A", "人物B", "关系类型", "共现章数"]
    _write_header(ws, headers)

    edges = relations.get("edges", [])
    sorted_edges = sorted(edges, key=lambda e: e.get("weight", 0), reverse=True)

    for edge in sorted_edges[:50]:
        ws.append([
            edge.get("source", ""),
            edge.get("target", ""),
            edge.get("relation_type", ""),
            edge.get("weight", 0),
        ])

    _auto_width(ws, headers)


def _render_locations(wb: Workbook, locations: list[dict]) -> None:
    ws = wb.create_sheet("地点表")
    headers = ["名称", "类型", "上级", "下级地点", "描述", "到访者", "首次出现", "提及章数"]
    _write_header(ws, headers)

    for loc in locations:
        children = loc.get("children", [])
        descriptions = loc.get("descriptions", [])
        desc_str = "; ".join(d.get("description", "")[:50] for d in descriptions[:2])
        visitors = loc.get("visitors", [])
        visitor_str = ", ".join(v["name"] for v in visitors[:6])
        stats = loc.get("stats", {})

        ws.append([
            loc["name"],
            loc.get("location_type", ""),
            loc.get("parent", ""),
            ", ".join(children[:8]),
            desc_str,
            visitor_str,
            stats.get("first_chapter", ""),
            stats.get("chapter_count", 0),
        ])

    _auto_width(ws, headers)


def _render_items(wb: Workbook, items: list[dict]) -> None:
    ws = wb.create_sheet("物品表")
    headers = ["名称", "类型", "流转记录"]
    _write_header(ws, headers)

    for item in items:
        flow = item.get("flow", [])
        flow_str = "; ".join(
            f"Ch.{f.get('chapter', '')} {f.get('actor', '')} {f.get('action', '')}"
            for f in flow[:5]
        )
        ws.append([
            item["name"],
            item.get("item_type", ""),
            flow_str,
        ])

    _auto_width(ws, headers)


def _render_orgs(wb: Workbook, orgs: list[dict]) -> None:
    ws = wb.create_sheet("组织表")
    headers = ["名称", "类型", "成员变动", "组织关系"]
    _write_header(ws, headers)

    for org in orgs:
        members = org.get("member_events", [])
        member_str = "; ".join(
            f"Ch.{m.get('chapter', '')} {m.get('member', '')}({m.get('action', '')})"
            for m in members[:5]
        )
        org_rels = org.get("org_relations", [])
        rel_str = ", ".join(
            f"{r.get('other_org', '')}({r.get('relation_type', '')})"
            for r in org_rels[:3]
        )
        ws.append([
            org["name"],
            org.get("org_type", ""),
            member_str,
            rel_str,
        ])

    _auto_width(ws, headers)


def _render_timeline(wb: Workbook, events: list[dict]) -> None:
    ws = wb.create_sheet("时间线")
    headers = ["章节", "事件摘要", "类型", "重要度", "参与者", "地点"]
    _write_header(ws, headers)

    for ev in events[:1000]:
        participants = ev.get("participants", [])
        ws.append([
            ev.get("chapter", 0),
            ev.get("summary", ""),
            ev.get("type", ""),
            ev.get("importance", ""),
            ", ".join(participants[:4]),
            ev.get("location", "") or "",
        ])

    _auto_width(ws, headers)


# ── Helpers ────────────────────────────────────────


def _write_header(ws, headers: list[str]) -> None:
    """Write styled header row."""
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.freeze_panes = "A2"


def _auto_width(ws, headers: list[str]) -> None:
    """Auto-adjust column widths based on content."""
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    cell.alignment = _CELL_ALIGN
                    val_len = len(str(cell.value))
                    if val_len > max_len:
                        max_len = val_len
        # Cap width at 50 characters, min 10
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)
